import openpyxl
import random
from openpyxl.styles import PatternFill
import shutil

def main():
    print("start")
    shutil.copy('./シフト表.xlsx', './完成版.xlsx')
    wb = openpyxl.load_workbook('./完成版.xlsx')
    ws = wb['work']
    shift_create(ws)
    wb.save('./完成版.xlsx')
    wb.close()
    print("end")

# シフト表にランダムでシフト記号を記入
def shift_create(ws):
    # 勤務体系（正規）
    shift_list_regular = ["A", "B", "C", "D", "E", "800"]
    # 勤務体系（非正規）
    shift_list_no_regular = ["F", "G", "H","730", "800", "1130"]
    
    early_turn = ["A", "B", "F", "730", "800"]
    late_turn = ["D", "E", "H", "1130"]
    
    fill_early = PatternFill(patternType='solid', fgColor='a9ceec')
    fill_late = PatternFill(patternType='solid', fgColor='7cfc00')
    fill_normal = PatternFill(patternType='solid', fgColor='ffffff')
    second_list = [4,7,10,13,16,19,22,25]
    third_list = [5,8,11,14,17,20,23,26]

    for raw in range(3, 27):
        shift_list = []
        for column in range(4, 26):
            shift_value = None
            while True:
                if ws.cell(raw, 3).value == "正規":
                    shift_value = random.choice(shift_list_regular)
                else:
                    shift_value = random.choice(shift_list_no_regular)

                if shift_list.count("A") == 3 and shift_value == "A":
                    continue
                if shift_list.count("E") == 3 and shift_value == "E":
                    continue
                if shift_list.count("F") == 3 and shift_value == "A":
                    continue
                if shift_list.count("H") == 3 and shift_value == "E":
                    continue
                if raw in second_list:
                    value1 = ws.cell(raw - 1, column).value
                    if (shift_value in early_turn) and value1 in early_turn:
                        continue
                    if (shift_value in late_turn) and value1 in late_turn:
                        continue
                if raw in third_list:
                    value1 = ws.cell(raw - 1, column).value
                    value2 = ws.cell(raw - 2, column).value
                    if (shift_value in early_turn):
                        if (value1 in early_turn):
                            continue
                        if (value2 in early_turn):
                            continue
                    if (shift_value in late_turn):
                        if (value1 in late_turn):
                            continue
                        if (value2 in late_turn):
                            continue
                    if shift_value not in early_turn and value1 not in early_turn and value2 not in early_turn:
                        continue
                    if shift_value not in late_turn:
                        if (value1 not in late_turn and early_turn) and (value2 not in late_turn and early_turn):
                            j = random.randint(1,2)
                            ws.cell(raw - j, column).value = "D" if random.random() >= 0.5 else "1130"
                            ws.cell(raw - j, column).fill = fill_late
                            continue
                if value == "A" and ws.cell(raw, column - 1).value == "E":
                    continue
                if value == "F" and ws.cell(raw, column - 1).value == "H":
                    continue
                break

            if shift_value in early_turn:
                ws.cell(raw, column).fill = fill_early
            elif shift_value in late_turn:
                ws.cell(raw, column).fill = fill_late
            else:
                ws.cell(raw, column).fill = fill_normal

            ws.cell(raw, column).value = shift_value
            shift_list.append(shift_value)
        print(shift_list)

if __name__ == "__main__":
    main()
